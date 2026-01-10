import customtkinter as ctk
from tkinter import messagebox, filedialog
import uuid
import threading
import datetime
import os
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from modules.assistants.chat_service import ChatService

class ChatFrame(ctk.CTkFrame):
    def __init__(self, master, app, assistant_data):
        super().__init__(master, fg_color="transparent")
        self.app = app
        self.assistant = assistant_data
        self.assistant_id = str(self.assistant.get('id', 'default'))
        
        # Initialize Service
        self.chat_service = ChatService(app.data_manager, self.assistant)
        
        self.current_conversation_id = None
        self.current_title = "Nouvelle conversation"
        self.history = [] # Mesage history for the current session
        
        # Header avec bouton retour
        header_frame = ctk.CTkFrame(self, fg_color="transparent", height=60)
        header_frame.pack(fill="x", padx=20, pady=10)
        header_frame.pack_propagate(False)
        
        btn_back = ctk.CTkButton(
            header_frame,
            text="< Retour",
            width=100,
            height=32,
            fg_color=("#3B8ED0", "#1F6AA5"),
            corner_radius=16,
            command=self.app.show_list,
        )
        btn_back.pack(side="left")
        
        title = ctk.CTkLabel(
            header_frame,
            text=f"💬 Chat avec {self.assistant.get('name', 'Assistant')}",
            font=("Arial", 20, "bold")
        )
        title.pack(side="left", padx=20)
        
        # Bouton Export Excel
        btn_export = ctk.CTkButton(
            header_frame,
            text="📥 Export Excel",
            width=120,
            height=32,
            fg_color=("#2E7D32", "#1B5E20"), # Vert foncé
            corner_radius=16,
            command=self.export_to_excel
        )
        # Fix: Helper to call export with current history if service doesn't manage state fully yet
        # Actually in my service implementation I copied export_to_excel but it uses self.history
        # The service I created DOES NOT have self.history.
        # I need to fix this. The View (ChatFrame) holds the state (history). Service should be stateless or I sync state.
        # Ideally Service should execute logic.
        # For now, I will keep export_to_excel in View OR pass history to Service.
        # Checking service code... I included export_to_excel but it relies on self.history which is empty/undefined in service.
        # I will keep export_to_excel in ChatFrame for now or fix Service to accept history.
        # Let's keep export_to_excel in View for this refactor to avoid breaking changes, as it interacts with UI (filedialog).
        # Wait, the service implementation I wrote in step 68 has export_to_excel which uses `self.history`. This is a bug in my service.
        # I'll Fix it in View.
        
        btn_export.configure(command=self.export_to_excel)
        btn_export.pack(side="right", padx=10)
        
        # Indicateur de provider
        provider_label = ctk.CTkLabel(
            header_frame,
            text=f"🤖 {self.assistant.get('provider', 'Non défini')}",
            font=("Arial", 12),
            text_color="gray"
        )
        provider_label.pack(side="right", padx=10)
        
        # --- Multi-Session Sidebar Setup ---
        self._setup_layout()
        
        # Initialiser l'UI
        self.after(100, self.refresh_history_list)
        
        # Message de bienvenue et tests de connexion
        self.add_system_message(f"Connexion à {self.assistant.get('name')}...")
        if self.assistant.get('description'):
            self.add_system_message(f"Description : {self.assistant.get('description')}")
        

        # Tester les connexions LLM
        self.after(100, self.test_llm_connections)
        
        # Envoyer automatiquement un message de bienvenue au LLM après les tests si l'historique est vide
        if not self.history:
            self.after(2000, self.send_welcome_message)

    def _setup_layout(self):
        """Configure le layout principal avec sidebar et zone de chat."""
        # Créer un container pour le corps de la page
        self.body_container = ctk.CTkFrame(self, fg_color="transparent")
        self.body_container.pack(fill="both", expand=True, padx=20, pady=(0, 20))
        
        # Configuration des colonnes : Sidebar (200px) | Chat (Espace restant)
        self.body_container.grid_columnconfigure(0, weight=0, minsize=220)
        self.body_container.grid_columnconfigure(1, weight=1)
        self.body_container.grid_rowconfigure(0, weight=1)
        
        # --- Sidebar Historique ---
        self.sidebar_history = ctk.CTkFrame(self.body_container, width=220, corner_radius=15)
        self.sidebar_history.grid(row=0, column=0, sticky="nsew", padx=(0, 15))
        self.sidebar_history.grid_propagate(False)
        
        lbl_hist = ctk.CTkLabel(self.sidebar_history, text="📜 Historique", font=("Arial", 15, "bold"))
        lbl_hist.pack(pady=15)
        
        self.btn_new_chat = ctk.CTkButton(
            self.sidebar_history,
            text="+ Nouvelle session",
            fg_color=("#4CAF50", "#388E3C"),
            hover_color=("#45A049", "#2E7D32"),
            height=35,
            corner_radius=10,
            command=self.start_new_conversation
        )
        self.btn_new_chat.pack(pady=(0, 15), padx=15, fill="x")
        
        # Liste scrollable des conversations
        self.history_list = ctk.CTkScrollableFrame(self.sidebar_history, fg_color="transparent")
        self.history_list.pack(fill="both", expand=True, padx=5, pady=5)
        
        # --- Zone de Chat Principale ---
        self.main_chat_panel = ctk.CTkFrame(self.body_container, fg_color="transparent")
        self.main_chat_panel.grid(row=0, column=1, sticky="nsew")
        
        # Initialisation des widgets dans le panel principal
        self.chat_area = ctk.CTkTextbox(
            self.main_chat_panel,
            font=("Arial", 13),
            wrap="word"
        )
        self.chat_area.pack(fill="both", expand=True, pady=(0, 10))
        self.chat_area.configure(state="disabled")
        
        self.progress_bar = ctk.CTkProgressBar(
            self.main_chat_panel,
            mode="indeterminate",
            height=8,
            corner_radius=4,
            progress_color=("#4CAF50", "#4CAF50"),
            fg_color=("gray85", "gray25")
        )
        self.progress_bar.set(0)
        
        self.input_frame = ctk.CTkFrame(self.main_chat_panel, fg_color="transparent")
        self.input_frame.pack(fill="x")
        self.input_frame.grid_columnconfigure(0, weight=1)
        
        self.entry = ctk.CTkEntry(
            self.input_frame,
            placeholder_text="Tapez votre message...",
            height=50,
            font=("Arial", 13)
        )
        self.entry.grid(row=0, column=0, sticky="ew", padx=(0, 10))
        self.entry.bind("<Return>", lambda e: self.send_message())
        
        self.btn_send = ctk.CTkButton(
            self.input_frame,
            text="Envoyer",
            width=100,
            height=50,
            corner_radius=25,
            fg_color=("#4CAF50", "#388E3C"),
            hover_color=("#45A049", "#2E7D32"),
            font=("Arial", 13, "bold"),
            command=self.send_message
        )
        self.btn_send.grid(row=0, column=1)

    # --- Conversation Management UI ---

    def start_new_conversation(self):
        """Réinitialise le tchat pour une nouvelle conversation."""
        self.current_conversation_id = None
        self.current_title = "Nouvelle conversation"
        self.history = []
        
        self.chat_area.configure(state="normal")
        self.chat_area.delete("1.0", "end")
        self.chat_area.configure(state="disabled")
        
        self.add_system_message(f"Connexion à {self.assistant.get('name')}...")
        if self.assistant.get('description'):
            self.add_system_message(f"Description : {self.assistant.get('description')}")
            
        self.refresh_history_list()
        
    def refresh_history_list(self):
        """Mise à jour de la liste des conversations dans la sidebar."""
        # Nettoyer la liste
        for widget in self.history_list.winfo_children():
            widget.destroy()
            
        conversations = self.app.data_manager.get_assistant_conversations("assistants", self.assistant_id)
        
        # Trier par date (la plus récente en haut)
        try:
            conversations.sort(key=lambda x: x.get('updated_at', ''), reverse=True)
        except:
            pass

        for conv in conversations:
            f = ctk.CTkFrame(self.history_list, fg_color="transparent")
            f.pack(fill="x", pady=2)
            
            is_active = (self.current_conversation_id == conv["id"])
            bg_color = ("gray80", "gray30") if is_active else "transparent"
            text_weight = "bold" if is_active else "normal"

            # Bouton de sélection
            btn = ctk.CTkButton(
                f,
                text=conv.get("title", "Sans titre"),
                anchor="w",
                fg_color=bg_color,
                text_color=("black", "white"),
                hover_color=("gray75", "gray35"),
                font=("Arial", 12, text_weight),
                height=32,
                corner_radius=8,
                command=lambda c=conv: self.load_conversation(c)
            )
            btn.pack(side="left", fill="x", expand=True, padx=(0, 5))

            # Bouton Renommer
            rename_btn = ctk.CTkButton(
                f,
                text="✎",
                width=28,
                height=28,
                fg_color="transparent",
                text_color="gray",
                hover_color=("gray90", "gray40"),
                command=lambda cid=conv["id"]: self.rename_conversation_ui(cid)
            )
            rename_btn.pack(side="right")
            
            # Bouton Supprimer
            del_btn = ctk.CTkButton(
                f,
                text="X",
                width=28,
                height=28,
                fg_color="transparent",
                text_color="#F44336",
                hover_color=("mistyrose", "darkred"),
                command=lambda cid=conv["id"]: self.delete_conversation_ui(cid)
            )
            del_btn.pack(side="right")

    def load_conversation(self, conversation):
        """Charge une session existante."""
        self.current_conversation_id = conversation["id"]
        self.current_title = conversation.get("title", "Sans titre")
        self.history = conversation.get("messages", [])
        
        self.chat_area.configure(state="normal")
        self.chat_area.delete("1.0", "end")
        
        for msg in self.history:
            role = msg.get("role", "Inconnu")
            content = msg.get("content", "")
            
            if role == "Système":
                self.chat_area.insert("end", f"ℹ️ {content}\n\n", "system")
            elif role == "Utilisateur":
                self.chat_area.insert("end", f"Vous : {content}\n\n", "user")
            elif role == "Assistant":
                name = self.assistant.get('name', 'Assistant')
                self.chat_area.insert("end", f"{name} : {content}\n\n", "assistant")
            elif role == "Erreur":
                self.chat_area.insert("end", f"❌ Erreur : {content}\n\n", "error")
        
        self.chat_area.configure(state="disabled")
        self.chat_area.see("end")
        self.refresh_history_list()

    def rename_conversation_ui(self, conversation_id):
        dialog = ctk.CTkInputDialog(text="Nouveau nom de session :", title="Renommer")
        new_title = dialog.get_input()
        if new_title:
            self.app.data_manager.rename_assistant_conversation("assistants", self.assistant_id, conversation_id, new_title)
            if self.current_conversation_id == conversation_id:
                self.current_title = new_title
            self.refresh_history_list()

    def delete_conversation_ui(self, conversation_id):
        if messagebox.askyesno("Confirmer", "Supprimer cette session d'historique ?"):
            self.app.data_manager.delete_assistant_conversation("assistants", self.assistant_id, conversation_id)
            if self.current_conversation_id == conversation_id:
                self.start_new_conversation()
            else:
                self.refresh_history_list()

    def _save_current_state(self):
        """Sauvegarde l'état actuel de la conversation."""
        if not self.current_conversation_id:
            self.current_conversation_id = str(uuid.uuid4())
            
        # Générer un titre si c'est le début
        if self.current_title == "Nouvelle conversation" and len(self.history) > 0:
            first_user_msg = next((m['content'] for m in self.history if m['role'] == 'Utilisateur'), None)
            if first_user_msg:
                self.current_title = " ".join(first_user_msg.split()[:5]) + "..."

        conversation = {
            "id": self.current_conversation_id,
            "title": self.current_title,
            "updated_at": str(datetime.datetime.now()),
            "messages": self.history
        }
        
        self.app.data_manager.save_assistant_conversation("assistants", self.assistant_id, conversation)
        self.refresh_history_list()
    
    def test_llm_connections(self):
        """Teste les connexions via le service."""
        # Use a lambda to be able to schedule UI updates from thread or just call direct since tests are usually fast or we can thread it
        # Original code didn't thread the test launch, but the tests themselves were synchronous.
        # But wait, original code did: self.after(100, self.test_llm_connections)
        # We can run this in a thread to be safe.
        
        def run_tests():
            # Define callback for logging
            def log_callback(msg):
                 self.after(0, lambda: self.add_system_message(msg))
            
            self.chat_service.test_connections(log_callback=log_callback)
            
        threading.Thread(target=run_tests, daemon=True).start()

    def send_welcome_message(self):
        """Envoie automatiquement un message de bienvenue au LLM."""
        if self.assistant.get('target_url'):
            welcome_msg = "Bonjour ! Présente-toi brièvement et lance immédiatement la recherche sur le site cible en fonction de ton objectif. IMPORTANT : Respecte scrupuleusement les consignes définies dans tes instructions (Contexte, Objectif, Limites)."
        else:
            welcome_msg = "Bonjour ! Peux-tu te présenter brièvement ?"
            
        self.add_user_message(welcome_msg)
        self.show_loading()
        self.btn_send.configure(state="disabled", text="Envoi...")
        
        threading.Thread(target=self._send_to_llm, args=(welcome_msg,)).start()
    
    def show_loading(self):
        self.progress_bar.pack(fill="x", padx=20, pady=(0, 10), before=self.input_frame)
        self.progress_bar.start()
    
    def hide_loading(self):
        self.progress_bar.stop()
        self.progress_bar.pack_forget()
    
    def add_system_message(self, text):
        if not self.winfo_exists() or not hasattr(self, 'chat_area') or not self.chat_area.winfo_exists():
            return
        self.history.append({"role": "Système", "content": text, "timestamp": str(datetime.datetime.now())})
        self.chat_area.configure(state="normal")
        self.chat_area.insert("end", f"ℹ️ {text}\n\n", "system")
        self.chat_area.tag_config("system", foreground="gray")
        self.chat_area.configure(state="disabled")
        self.chat_area.see("end")
        self._save_current_state()
    
    def add_user_message(self, text):
        if not self.winfo_exists() or not hasattr(self, 'chat_area') or not self.chat_area.winfo_exists():
            return
        self.history.append({"role": "Utilisateur", "content": text, "timestamp": str(datetime.datetime.now())})
        self.chat_area.configure(state="normal")
        self.chat_area.insert("end", f"Vous : {text}\n\n", "user")
        self.chat_area.tag_config("user", foreground="#2196F3")
        self.chat_area.configure(state="disabled")
        self.chat_area.see("end")
        self._save_current_state()
    
    def add_assistant_message(self, text):
        if not self.winfo_exists() or not hasattr(self, 'chat_area') or not self.chat_area.winfo_exists():
            return
        self.history.append({"role": "Assistant", "content": text, "timestamp": str(datetime.datetime.now())})
        self.chat_area.configure(state="normal")
        self.chat_area.insert("end", f"{self.assistant.get('name')} : {text}\n\n", "assistant")
        self.chat_area.tag_config("assistant", foreground="#4CAF50")
        self.chat_area.configure(state="disabled")
        self.chat_area.see("end")
        self._save_current_state()
    
    def add_error_message(self, text):
        if not self.winfo_exists() or not hasattr(self, 'chat_area') or not self.chat_area.winfo_exists():
            return
        self.history.append({"role": "Erreur", "content": text, "timestamp": str(datetime.datetime.now())})
        self.chat_area.configure(state="normal")
        self.chat_area.insert("end", f"❌ Erreur : {text}\n\n", "error")
        self.chat_area.tag_config("error", foreground="#F44336")
        self.chat_area.configure(state="disabled")
        self.chat_area.see("end")
        self._save_current_state()
    
    def build_system_prompt(self):
        """Construit le prompt système avec toutes les informations de l'assistant."""
        # This logic is purely string formatting based on config.
        # We can keep it here or move it to service. 
        # Moving to service is better to be consistent.
        # But for now, let's duplicate logic or just keep it here.
        # Ideally, `ChatService` should offer `_get_system_prompt`.
        # I'll keep it here for now to minimize changes in Logic, but using service for LLM calls.
        # Wait, I am delegating to service. generate_response expects system_prompt.
        
        effective_config = self.app.data_manager.get_effective_assistant_config(self.assistant["id"])
        
        parts = []
        if effective_config.get('role'): parts.append(f"Rôle : {effective_config.get('role')}")
        if effective_config.get('context'): parts.append(f"Contexte : {effective_config.get('context')}")
        if effective_config.get('objective'): parts.append(f"Objectif : {effective_config.get('objective')}")
        if effective_config.get('limits'): parts.append(f"Limites : {effective_config.get('limits')}")
        if effective_config.get('response_format'): parts.append(f"Format de réponse : {effective_config.get('response_format')}")
            
        target_url = effective_config.get('target_url')
        if target_url:
            parts.append(f"IMPORTANT : Tu as accès à un outil de recherche sur le site : {target_url}\nPour effectuer une recherche sur ce site, réponds UNIQUEMENT avec la commande suivante :\nACTION: SEARCH <ta requête de recherche>\n\nExemple :\nUtilisateur : \"Cherche des chaussures rouges\"\nToi : ACTION: SEARCH chaussures rouges\n\nJe t'enverrai ensuite les résultats de la recherche, et tu pourras formuler ta réponse finale.\nN'utilise cette commande que si c'est pertinent pour répondre à l'utilisateur.")
            
            url_instructions = effective_config.get('url_instructions')
            if url_instructions:
                parts.append(f"INSTRUCTIONS POUR LE SITE {target_url} :\nLe système utilise l'IA pour extraire automatiquement les données selon ces instructions :\n{url_instructions}\n\nTu n'as pas besoin de gérer l'extraction toi-même - l'IA s'en charge.\nConcentre-toi sur la formulation de requêtes de recherche pertinentes et l'analyse des résultats retournés.")

        parts.append("IMPORTANT :\n1. Tu dois analyser et comprendre le fonctionnement du site internet cible pour naviguer et extraire les informations pertinentes.\n2. MAIS SURTOUT : Ta PRIORITÉ ABSOLUE est de respecter scrupuleusement les consignes définies ci-dessus (Rôle, Contexte, Objectif, Limites).\n3. En cas de conflit entre une information du site et tes instructions, tes instructions (Limites notamment) prévalent toujours.")

        return "\n\n".join(parts) if parts else "Tu es un assistant utile et serviable."
    
    def send_message(self):
        """Envoie un message au LLM."""
        user_message = self.entry.get().strip()
        if not user_message: return
        
        self.add_user_message(user_message)
        self.entry.delete(0, "end")
        self.show_loading()
        self.btn_send.configure(state="disabled", text="Envoi...")
        
        threading.Thread(target=self._send_to_llm, args=(user_message,), daemon=True).start()
    
    def _send_to_llm(self, user_message):
        """Envoie la requête au LLM (dans un thread séparé) via le service."""
        try:
            system_prompt = self.build_system_prompt()
            
            # Step 1: Initial Generation
            # Callback for RAG or other system messages during generation
            def sys_callback(msg):
                self.after(0, lambda: self.add_system_message(msg))
                
            response = self.chat_service.generate_response(
                user_message, system_prompt, system_msg_callback=sys_callback
            )
            
            if not response['success']:
                self.after(0, lambda: self.add_error_message(response['error']))
            else:
                response_text = response['text']
                api_key = response['api_key']
                
                # Step 2: Process response actions (Scraping, etc.) using generator
                # The generator yields status updates and finally the result
                
                def process_generator():
                    generator = self.chat_service.process_response_action(
                        response_text, api_key, system_prompt, user_message, system_msg_callback=sys_callback
                    )
                    
                    final_content = None
                    try:
                        for item in generator:
                            if item:
                                if item['type'] == 'text':
                                    final_content = item['content']
                                    # If it's a chunk of text, we can display it? 
                                    # My service implementation yields 'text' for final result, or intro.
                                    # If intro, display as assistant message.
                                    # If final, we wait for end?
                                    # Let's handle it:
                                    self.after(0, lambda c=item['content']: self.add_assistant_message(c))
                                elif item['type'] == 'error':
                                    self.after(0, lambda c=item['content']: self.add_error_message(c))
                    except Exception as gen_e:
                         self.after(0, lambda: self.add_error_message(str(gen_e)))
                         
                # Execute generator logic synchronously in this thread
                process_generator()
                
        except Exception as e:
            self.after(0, lambda: self.add_error_message(f"❌ Erreur technique : {str(e)}"))
        
        finally:
            self.after(0, self._finalize_llm_call)

    def _finalize_llm_call(self):
        if not self.winfo_exists(): return
        self.hide_loading()
        if hasattr(self, 'btn_send') and self.btn_send.winfo_exists():
            self.btn_send.configure(state="normal", text="Envoyer")

    def export_to_excel(self):
        """
        Copie locale de export_to_excel car le service est stateless.
        On garde cette méthode ici car elle dépend fortement de self.history.
        """
        # ... (Code copied from original ChatFrame export_to_excel) ...
        # Since I'm using write_to_file, I need to include the full body of export_to_excel
        if not self.history:
            messagebox.showinfo("Info", "Aucun message à exporter.")
            return

        target_section = "Partie 2 : Synthèse à exporter"
        found_tables = []
        
        for i, msg in enumerate(self.history):
            if msg["role"] == "Assistant" and target_section in msg["content"]:
                content = msg["content"]
                start_index = content.find(target_section) + len(target_section)
                section_content = content[start_index:]
                # We need _parse_markdown_table which was also in ChatFrame.
                # I should probably move helper methods to a utility or keep them here.
                # Since I moved export_to_excel to ChatService in my previous thought but realized it was broken,
                # I will define _parse_markdown_table here as helper.
                table_data = self._parse_markdown_table(section_content)
                if table_data:
                    timestamp = msg.get("timestamp", datetime.datetime.now())
                    # Convert string timestamp back to obj if needed or just use
                    ts_str = str(timestamp) # Simplified
                    found_tables.append({
                        "id": len(found_tables) + 1, "timestamp": ts_str, "data": table_data
                    })
        
        if not found_tables:
            messagebox.showwarning("Attention", f"Aucune table trouvée dans les sections '{target_section}'.")
            return
            
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"chatbot_export_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                title=f"Exporter {len(found_tables)} résultat(s)"
            )
            if not filename: return
            
            wb = Workbook()
            default_ws = wb.active
            header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            for index, item in enumerate(found_tables):
                if index == 0: ws = default_ws
                else: ws = wb.create_sheet()
                sheet_title = f"Resultat {item['id']}"
                ws.title = sheet_title
                table_data = item["data"]
                if table_data["headers"]:
                    ws.append(table_data["headers"])
                    for cell in ws[1]:
                        cell.fill = header_fill; cell.font = header_font; cell.alignment = center_align; cell.border = thin_border
                for row in table_data["rows"]:
                    ws.append(row)
                    for cell in ws[ws.max_row]: cell.border = thin_border; cell.alignment = Alignment(vertical="center", wrap_text=True)
                for col in ws.columns:
                    max_len = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column].width = min(max_len + 2, 60)
            wb.save(filename)
            messagebox.showinfo("Succès", f"Export terminé avec succès !")
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export Excel : {str(e)}")

    def _parse_markdown_table(self, text):
        lines = text.strip().split('\n')
        headers = []; rows = []; in_table = False
        for i, line in enumerate(lines):
            line = line.strip()
            if "|" in line:
                parts = [p.strip() for p in line.split('|') if p.strip()]
                if not parts: continue
                if not in_table:
                    if i + 1 < len(lines) and "---" in lines[i+1]:
                        headers = parts; in_table = True; continue
                elif "---" in line: continue
                else: rows.append(parts)
            elif in_table and not line: break
        if headers or rows: return {"headers": headers, "rows": rows}
        return None
